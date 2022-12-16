import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showwarning, showinfo

import openpyxl


class Excel:
    def __init__(self, dataT, dataE, cell_name, count):
        self.dataT = ""
        self.dataE = ""
        self.cell_name = ""
        self.count = 0


def main():
    window = tk.Tk()
    window.title("Excel reader script")
    window.geometry("200x200")
    window.resizable(width=False, height=False)
    canvas = tk.Canvas(window, height=400, width=300)
    canvas.pack()
    frame = tk.Frame(window)
    frame.place(relheight=1, relwidth=1)
    Excel.count = 0
    btn_excelFile = tk.Button(frame, text="Выбери excel", command=file_read_e)
    btn_excelFile.pack(anchor="n", padx=10, pady=5)
    btn_txtFile = tk.Button(frame, text="Выбери txt файл", command=file_read_t)
    btn_txtFile.pack(anchor="n", padx=10, pady=10)
    lbl_name = tk.Label(frame, text="Введи номер ячейки:")
    lbl_name.pack(anchor="n", padx=10, pady=2)
    Excel.cell_name = tk.Entry(frame)
    Excel.cell_name.pack(anchor="n", padx=10, pady=2)
    btn_confirm = tk.Button(frame, text="Выполнить", command=write_adds)
    btn_confirm.pack(anchor="n", padx=10, pady=15)

    window.mainloop()


def file_read_e():
    Excel.dataE = askopenfilename(defaultextension="xlsx")
    if not Excel.dataE or not "*.xlsx":
        showwarning("Предупреждение", "Не выбрана таблица")


def file_read_t():
    Excel.dataT = askopenfilename(defaultextension="txt")
    Excel.count = 1
    if not Excel.dataT or not "*.txt":
        showwarning("Предупреждение", "Не выбран текстовый файл")


def write_adds():
    ent_cell = Excel.cell_name.get()
    if Excel.count == 1:
        write_file_adds(ent_cell)
    else:
        write_oak(ent_cell)


def load_workbook():
    path = Excel.dataE
    wb = openpyxl.load_workbook(path)
    sheets = wb.worksheets
    return path, wb, sheets


def write_oak(ent_cell):
    path, wb, sheets = load_workbook()
    for sheet, i in zip(
            sheets, range(150)):
        sheet[f"{ent_cell}"] = None
        sheet[f"{ent_cell}"] = f"#{i + 1}/АООК инв. 50232"
    wb.save(path)
    showinfo("Выполнено", "Листы пронумерованы AООК")


def write_file_adds(ent_cell):
    path, wb, sheets = load_workbook()
    data = []
    with open(Excel.dataT, "r+", encoding="utf-8") as f:
        data = [line.rstrip('\n') for line in f]
    f.close()
    if len(data) > 2:
        for sheet, i in zip(
                sheets, data):
            sheet[f"{ent_cell}"] = None
            sheet[f"{ent_cell}"] = i
        wb.save(path)
        showinfo("Выполнено", "Данные записаны в файл")
    else:
        showwarning("Ошибка", "Пустой файл")
    Excel.count = 0
    Excel.dataT = None


if __name__ == "__main__":
    main()
